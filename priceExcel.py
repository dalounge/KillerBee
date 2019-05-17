import csv
from openpyxl import Workbook, load_workbook

# Path to excel sheet.  Use a r'' string [raw] if you don't want to use \\
path = 'c:\\PriceListReader\\Copy of v2017 Wonderware Price List Update for CS 1_2018.xlsx'

# Load excel sheet into memory.  Remember this is the memory version of the excel sheet
wb = load_workbook(path)

# Load up the sheet we want to work with
wbHist = wb['Historian 2017']

get_values = False
get_indexes = False
prices = {}
listPrice_idx = None
partNo_idx = None
partNo_name = None

# This iterates through each row
for row_cells in wbHist.iter_rows():
    # Once we find a table with part #s, we start grabbing the list price
    if get_indexes:
        get_indexes = False
        get_values = True
    
    # This now grabs each cell in each row
    for cells in row_cells:
        # Starts the index once we find Part description for the table
        if cells.value == 'Part Description':
            get_indexes = True
            continue
        # Indexes which column th part number is
        if get_indexes:
            if cells.value == 'Part Number':
                partNo_idx = cells.column
                continue
        # Indexes which column the part number is
            elif cells.value == 'List Price':
                listPrice_idx = cells.column
                continue
        # With the index above, grabs the part number
        if get_values:
            if cells.column == partNo_idx:
                prices[cells.value] = {}
                partNo_name = cells.value
                continue
            # Grabs the list number and stores the location.  Use location to modify price
            elif cells.column == listPrice_idx:
                prices[partNo_name]['ListPrice'] = cells.value
                prices[partNo_name]['Location'] = [cells.column, cells.row]
                continue
            # If it sees a blank it will reset for us to find the next table
            # This needs to be improved in case something is written
            # Start reading additional columns for blanks such as list price to confirm
            elif cells.value == '':
                get_values = False
                get_indexes = False
                listPrice_idx = None
                partNo_idx = None
                partNo_name = None
                break
                
print(prices)
